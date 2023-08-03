import pandas as pd
import numpy as np
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

COLOR_VALID = "CCFFCC"
COLOR_INVALID = "FF8080"
COLOR_DEFAULT = "FBCEB1"
COLOR_NO_OFFICE = "BEBEBE"

NETWORK_SHARE_DIRECTORY = r'\\network_share'
DIRECTORY = Path(NETWORK_SHARE_DIRECTORY) / 'inventory' / 'csv'
OFFICE_FILENAME = Path(NETWORK_SHARE_DIRECTORY) / 'inventory' / 'office.xlsx'
REPORT_FILENAME = Path(NETWORK_SHARE_DIRECTORY) / 'inventory' / 'raport.xlsx'


def auto_format_column_width(worksheet: Worksheet) -> None:
    for column in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = max_length
        column_letter = get_column_letter(column[0].column)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def process_csv_files(csv_directory: Path) -> pd.DataFrame:
    def read_file(filepath: Path) -> pd.DataFrame:
        df = pd.read_csv(filepath, na_values=[np.nan], keep_default_na=False)  # Change NA values to "NA" string
        df.replace("", "-", inplace=True)  # Change empty values to "-"
        return df

    csv_files = [file for file in csv_directory.iterdir() if file.suffix == ".csv"]
    df = pd.concat([read_file(file) for file in csv_files], ignore_index=True)  # Use pd.concat for concatenating DataFrames
    return df


def process_office_data(df: pd.DataFrame) -> pd.DataFrame:
    office_df = df[['PC_Name', 'IP', 'User', 'OS', 'Office', 'Office_Keys', 'Office_Key_ID']]
    return office_df


def create_workbook(office_df: pd.DataFrame) -> tuple[Workbook, Worksheet]:
    workbook = Workbook()
    worksheet = workbook.active

    headers = office_df.columns.tolist()
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in office_df.itertuples(index=False):
        worksheet.append(row)

    auto_format_column_width(worksheet)

    return workbook, worksheet


def apply_formatting(worksheet: Worksheet, office_df: pd.DataFrame) -> None:
    name_column_index = office_df.columns.get_loc('Office') + 1
    pattern = r"2007 \(\d+\.\d+\)"

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        office_version_cell = row[name_column_index - 1]
        office_key_cell = row[name_column_index]
        office_key_id_cell = row[name_column_index + 1]
        if re.match(pattern, str(office_version_cell.value)):
            for cell in row:
                cell.fill = PatternFill(start_color=COLOR_VALID, fill_type="solid")
        elif office_version_cell.value == "NA" or office_version_cell.value =="-":
            for cell in row:
                cell.fill = PatternFill(start_color=COLOR_NO_OFFICE, fill_type="solid")
        elif office_key_cell.value in ["System.Object[]", None, "","-"]:
            for cell in row:
                cell.fill = PatternFill(start_color=COLOR_INVALID, fill_type="solid")
        else:
            for cell in row:
                cell.fill = PatternFill(start_color=COLOR_DEFAULT, fill_type="solid")


def save_workbook(workbook: Workbook, filename: Path) -> None:
    workbook.save(filename)


def main():
    with pd.option_context("display.max_columns", None):
        df = process_csv_files(DIRECTORY)
        office_df = process_office_data(df)

        workbook, worksheet = create_workbook(office_df)
        apply_formatting(worksheet, office_df)
        save_workbook(workbook, OFFICE_FILENAME)
        with pd.ExcelWriter(REPORT_FILENAME, engine='openpyxl') as writer:
            df.to_excel(writer, index=None, header=True)


if __name__ == "__main__":
    main()
